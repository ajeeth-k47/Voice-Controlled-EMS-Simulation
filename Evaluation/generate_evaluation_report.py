"""
Compare baseline vs generated hand motions (peak joint angles only — no timing).

Folders (under project root):
  Evaluation/Baseline/   — e.g. I.xlsx … TR.xlsx (from OpenSim Motionfiles)
  Evaluation/Generated/ — e.g. index_finger_flexion_coords.xlsx (from simulations/*_coords.mot via mot_to_excel)

Run from project root:
  python Evaluation/generate_evaluation_report.py

Writes:
  Evaluation/report/evaluation_summary.csv
  Evaluation/report/evaluation_details.csv
  Evaluation/report/evaluation_report.md

Optional: Evaluation/motion_pairs.csv with columns label,baseline_filename,generated_filename
If missing, uses the built-in 15-movement name map below.
"""

from __future__ import annotations

import argparse
import csv
import os
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from io import StringIO
from typing import Dict, List, Tuple

import pandas as pd

# -----------------------------------------------------------------------------
# Paths
# -----------------------------------------------------------------------------

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_DEFAULT_EVAL_ROOT = _SCRIPT_DIR  # Evaluation/

# When baseline peak is tiny, avoid divide-by-zero in % error
_MIN_BASE_DEG = 1.0


# -----------------------------------------------------------------------------
# Load .xlsx or OpenSim .mot / .sto
# -----------------------------------------------------------------------------


def _load_xlsx(path: str) -> pd.DataFrame:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return pd.DataFrame()
    header = [str(c) if c is not None else "" for c in rows[0]]
    return pd.DataFrame(rows[1:], columns=header)


def _load_opensim_text(path: str) -> pd.DataFrame:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        lines = f.readlines()
    end = None
    for i, line in enumerate(lines):
        if line.strip().lower() == "endheader":
            end = i
            break
    if end is None:
        raise ValueError(f"No endheader in {path}")
    col_names = lines[end + 1].strip().split()
    data_text = "".join(lines[end + 2 :]).strip()
    if not data_text:
        raise ValueError(f"No data in {path}")
    return pd.read_csv(StringIO(data_text), sep=r"\s+", header=None, names=col_names, engine="python")


def load_table(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xlsx":
        return _load_xlsx(path)
    if ext == ".xls":
        return pd.read_excel(path)
    return _load_opensim_text(path)


def flex_columns(df: pd.DataFrame) -> List[str]:
    """Columns to compare: *_flex and wrist flexion."""
    out: List[str] = []
    seen = set()
    for c in df.columns:
        if c == "time":
            continue
        s = str(c)
        if s.endswith("_flex") or s == "flexion":
            if s not in seen:
                seen.add(s)
                out.append(s)
    return out


def peak_max(df: pd.DataFrame, cols: List[str]) -> Dict[str, float]:
    return {c: float(df[c].astype(float).max()) for c in cols if c in df.columns}


@dataclass
class PairSummary:
    label: str
    n_active_joints: int
    mae_peak_deg: float
    rmse_peak_deg: float
    mean_relative_peak_error_pct: float
    peak_similarity_pct: float


def compare_pair(
    label: str,
    baseline_path: str,
    generated_path: str,
    active_threshold_deg: float,
) -> Tuple[PairSummary, pd.DataFrame]:
    """
    For each overlapping flex DOF: compare max angle (peak) baseline vs generated.
    Active DOFs = baseline peak > active_threshold_deg (else top 3 baseline peaks).

    peak_similarity_pct = 100 - mean(relative error %), clipped to [0, 100]
    relative error % per joint = 100 * |gen - base| / max(base, 1°)
    """
    db = load_table(baseline_path)
    dg = load_table(generated_path)

    base_cols = flex_columns(db)
    gen_set = set(flex_columns(dg))
    cols = [c for c in base_cols if c in gen_set]
    if not cols:
        raise ValueError(f"No overlapping flex columns for '{label}'")

    pb = peak_max(db, cols)
    pg = peak_max(dg, cols)

    active = [c for c, v in pb.items() if v > active_threshold_deg]
    if not active:
        active = sorted(pb.keys(), key=lambda k: pb[k], reverse=True)[:3]

    rows = []
    errs: List[float] = []
    rel_pcts: List[float] = []
    for c in active:
        b, g = pb[c], pg[c]
        err = abs(g - b)
        errs.append(err)
        denom = max(float(b), _MIN_BASE_DEG)
        rp = 100.0 * err / denom
        rel_pcts.append(rp)
        rows.append(
            {
                "label": label,
                "coordinate": c,
                "baseline_peak_deg": b,
                "generated_peak_deg": g,
                "abs_peak_error_deg": err,
                "relative_peak_error_pct": rp,
            }
        )

    mae = sum(errs) / len(errs)
    rmse = (sum(e * e for e in errs) / len(errs)) ** 0.5
    mean_rel = sum(rel_pcts) / len(rel_pcts)
    sim = max(0.0, min(100.0, 100.0 - mean_rel))

    summary = PairSummary(
        label=label,
        n_active_joints=len(active),
        mae_peak_deg=float(mae),
        rmse_peak_deg=float(rmse),
        mean_relative_peak_error_pct=float(mean_rel),
        peak_similarity_pct=float(sim),
    )
    return summary, pd.DataFrame(rows)


# -----------------------------------------------------------------------------
# 15 movements: baseline file stem -> generated xlsx stem (no .xlsx)
# -----------------------------------------------------------------------------

DEFAULT_PAIRS: List[Tuple[str, str, str]] = [
    ("I", "I", "index_finger_flexion_coords"),
    ("IM", "IM", "index_+_middle_flexion_coords"),
    ("IMR", "IMR", "index_+_middle_+_ring_flexion_coords"),
    ("IMRL", "IMRL", "hand_closed_coords"),
    ("L", "L", "little_finger_flexion_coords"),
    ("M", "M", "middle_finger_flexion_coords"),
    ("MR", "MR", "middle_+_ring_flexion_coords"),
    ("MRL", "MRL", "middle_+_ring_+_little_flexion_coords"),
    ("R", "R", "ring_finger_flexion_coords"),
    ("RL", "RL", "ring_and_little_finger_flexion_coords"),
    ("T", "T", "thumb_flexion_coords"),
    ("TI", "TI", "thumb_+_index_flexion_coords"),
    ("TL", "TL", "thumb_+_little_flexion_coords"),
    ("TM", "TM", "thumb_+_middle_flexion_coords"),
    ("TR", "TR", "thumb_and_ring_flexion_coords"),
]


def load_pairs_csv(path: str, baseline_dir: str, generated_dir: str) -> List[Tuple[str, str, str]]:
    out: List[Tuple[str, str, str]] = []
    with open(path, "r", encoding="utf-8", newline="") as f:
        for row in csv.DictReader(f):
            label = row["label"].strip()
            bf = row["baseline_filename"].strip()
            gf = row["generated_filename"].strip()
            if not bf.lower().endswith(".xlsx"):
                bf += ".xlsx"
            if not gf.lower().endswith(".xlsx"):
                gf += ".xlsx"
            out.append((label, os.path.join(baseline_dir, bf), os.path.join(generated_dir, gf)))
    return out


def default_paths(baseline_dir: str, generated_dir: str) -> List[Tuple[str, str, str]]:
    return [
        (lab, os.path.join(baseline_dir, f"{b}.xlsx"), os.path.join(generated_dir, f"{g}.xlsx"))
        for lab, b, g in DEFAULT_PAIRS
    ]


def markdown_table(df: pd.DataFrame) -> str:
    if df.empty:
        return "_No successful pairs._"
    cols = [str(c) for c in df.columns]
    head = "| " + " | ".join(cols) + " |"
    sep = "| " + " | ".join(["---"] * len(cols)) + " |"
    body = ["| " + " | ".join(str(row[c]) for c in df.columns) + " |" for _, row in df.iterrows()]
    return "\n".join([head, sep] + body)


def write_report_md(path: str, summary_df: pd.DataFrame, skipped: List[str], ts: str) -> None:
    lines = [
        "# Baseline vs generated (peak angles only)",
        "",
        f"Generated: {ts} UTC",
        "",
        "## Summary",
        "",
        markdown_table(summary_df),
        "",
    ]
    if skipped:
        lines += ["## Skipped", "", "\n".join(f"- {s}" for s in skipped), ""]
    if not summary_df.empty:
        lines += [
            "## Overall (mean over movements)",
            "",
            f"- MAE peak (deg): {float(summary_df['mae_peak_deg'].mean()):.4f}",
            f"- RMSE peak (deg): {float(summary_df['rmse_peak_deg'].mean()):.4f}",
            f"- Peak similarity (%): {float(summary_df['peak_similarity_pct'].mean()):.2f}",
            "",
        ]
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def main() -> None:
    p = argparse.ArgumentParser(description="Baseline vs generated peak comparison.")
    p.add_argument(
        "--evaluation-dir",
        default=_DEFAULT_EVAL_ROOT,
        help="Folder with Baseline/ and Generated/ (default: Evaluation/ next to this script).",
    )
    p.add_argument("--output-dir", default="", help="Report folder (default: <evaluation-dir>/report).")
    p.add_argument("--pairs-csv", default="", help="Optional motion_pairs.csv inside evaluation-dir.")
    p.add_argument("--active-threshold-deg", type=float, default=5.0, help="Baseline peak > this = active DOF.")
    args = p.parse_args()

    eval_root = os.path.abspath(args.evaluation_dir)
    baseline_dir = os.path.join(eval_root, "Baseline")
    generated_dir = os.path.join(eval_root, "Generated")
    out_dir = os.path.abspath(args.output_dir) if args.output_dir else os.path.join(eval_root, "report")
    os.makedirs(out_dir, exist_ok=True)

    pairs_path = args.pairs_csv.strip()
    if pairs_path and not os.path.isabs(pairs_path):
        pairs_path = os.path.join(eval_root, pairs_path)
    if not pairs_path:
        pairs_path = os.path.join(eval_root, "motion_pairs.csv")

    if os.path.isfile(pairs_path):
        print(f"Pairs: {pairs_path}")
        pairs = load_pairs_csv(pairs_path, baseline_dir, generated_dir)
    else:
        print("Pairs: built-in 15 movements")
        pairs = default_paths(baseline_dir, generated_dir)

    summaries: List[dict] = []
    details_list: List[pd.DataFrame] = []
    skipped: List[str] = []

    for label, bpath, gpath in pairs:
        if not os.path.isfile(bpath):
            skipped.append(f"{label}: missing {bpath}")
            continue
        if not os.path.isfile(gpath):
            skipped.append(f"{label}: missing {gpath}")
            continue
        try:
            s, ddf = compare_pair(label, bpath, gpath, args.active_threshold_deg)
            summaries.append(
                {
                    "label": s.label,
                    "baseline_file": os.path.basename(bpath),
                    "generated_file": os.path.basename(gpath),
                    "n_active_joints": s.n_active_joints,
                    "mae_peak_deg": s.mae_peak_deg,
                    "rmse_peak_deg": s.rmse_peak_deg,
                    "mean_relative_peak_error_pct": s.mean_relative_peak_error_pct,
                    "peak_similarity_pct": s.peak_similarity_pct,
                }
            )
            details_list.append(ddf)
            print(
                f"[{label}] joints={s.n_active_joints} "
                f"MAE={s.mae_peak_deg:.2f}° similarity={s.peak_similarity_pct:.1f}%"
            )
        except Exception as e:
            skipped.append(f"{label}: {e}")
            print(f"[{label}] ERROR: {e}", file=sys.stderr)

    summary_df = pd.DataFrame(summaries)
    details_df = pd.concat(details_list, ignore_index=True) if details_list else pd.DataFrame()

    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H-%M-%SZ")
    sum_csv = os.path.join(out_dir, "evaluation_summary.csv")
    det_csv = os.path.join(out_dir, "evaluation_details.csv")
    md_path = os.path.join(out_dir, "evaluation_report.md")

    summary_df.to_csv(sum_csv, index=False)
    details_df.to_csv(det_csv, index=False)
    write_report_md(md_path, summary_df, skipped, ts)

    print(f"\nWrote:\n  {sum_csv}\n  {det_csv}\n  {md_path}")
    if skipped:
        print(f"\nSkipped: {len(skipped)}")


if __name__ == "__main__":
    main()
